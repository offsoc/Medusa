#!/usr/bin/env python
# -*- coding: utf-8 -*-
from Web.DatabaseHub import UserInfo,EmailInfo
from django.http import JsonResponse,FileResponse
from ClassCongregation import ErrorLog,randoms,GetTempFilePath,TemplatePath
from openpyxl import load_workbook
import json
import time
import ast
from Web.Workbench.LogRelated import UserOperationLogRecord,RequestLogRecord
"""upload_email_list
{
POST /api/upload_email_list/ HTTP/1.1
Content-Type: multipart/form-data; boundary=----WebKitFormBoundaryaFtQbWz7pBzNgCOv
token:XXXXXXXXXXXXXXXX
Another-Name:\u5403\u996d

------WebKitFormBoundaryaFtQbWz7pBzNgCOv
Content-Disposition: form-data; name="file"; filename="test.xlsx"
Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet

XXXXXXXXXXXXXXX
------WebKitFormBoundaryaFtQbWz7pBzNgCOv--
}
"""
def Upload(request):#上传表格，提取相关数据
    RequestLogRecord(request, request_api="upload_email_list")
    if request.method == "POST":
        try:
            Token = request.headers["token"]
            Uid = UserInfo().QueryUidWithToken(Token)  # 如果登录成功后就来查询UID
            if Uid != None:  # 查到了UID
                UserOperationLogRecord(request, request_api="upload_email_list", uid=Uid)  # 查询到了在计入
                SaveRoute=""
                PictureData = request.FILES.get('file', None)#获取文件数据
                AnotherName = PictureData.name#文件名称
                if 0<PictureData.size:#内容不能为空
                    SaveFileName=randoms().result(50)+str(int(time.time()))#重命名文件
                    SaveRoute=GetTempFilePath().Result()+SaveFileName+".xlsx"#获得保存路径
                    with open(SaveRoute, 'wb') as f:
                        for line in PictureData:
                            f.write(line)
                ReadExcel = load_workbook(SaveRoute) #读取上传的文件
                ExcelData = ReadExcel[ReadExcel.sheetnames[0]]  # 获取第一个sheet
                # 按行读取 工作表的内容
                Excel = {}  # 创建一个空字典,存储表格数据
                for row in [row for row in ExcelData.rows][1:]:#删除了第一行数据
                    # print(row[0].value, row[1].value)
                    Department = str(row[0].value).replace("\n", "")  # 部门
                    Value = str(row[1].value).replace("\n", "")  # 值
                    if Department != "None" and Value != "None":  # 过滤空值
                        if Department in Excel.keys():  # 判断部门是否在键中
                            Excel[Department].append(Value)
                        else:
                            Excel[Department] = [Value]
                Result=EmailInfo().Write(uid=Uid,email_list=str(Excel),another_name=AnotherName,project_key=randoms().result(20))#写入数据库
                if Result:
                    return JsonResponse({'message': "写入成功！", 'code': 200, })
                else:
                    return JsonResponse({'message': "写入失败！", 'code': 501, })
            else:
                return JsonResponse({'message': "小宝贝这是非法查询哦(๑•̀ㅂ•́)و✧", 'code': 403, })
        except Exception as e:
            ErrorLog().Write("Web_Email_EmailList_Upload(def)", e)
            return JsonResponse({'message': "出错了请看报错日志(๑•̀ㅂ•́)و✧", 'code': 169, })
    else:
        return JsonResponse({'message': '请使用Post请求', 'code': 500, })

def Download(request):#下载模版
    RequestLogRecord(request, request_api="download_email_list_template")
    if request.method == "GET":
        try:
            Template=TemplatePath().Result()+"EmailListTemplate.xlsx"#获取邮件地址
            TemplateFlow = open(Template, 'rb')
            Result=FileResponse(TemplateFlow)#把图片比特流复制给返回包
            Result['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            Result['Content-Disposition'] = 'form-data;filename="EmailListTemplate.xlsx"'
            return Result
        except Exception as e:
            ErrorLog().Write("Web_Email_EmailList_Download(def)", e)
            return JsonResponse({'message': '呐呐呐！莎酱被玩坏啦(>^ω^<)', 'code': 169, })
    else:
        return JsonResponse({'message': '请使用GET请求', 'code': 500, })




"""statistics_email_list_key
{

	"token": "xxxx"
}
"""
def Statistics(request):#统计邮件列表个数数据
    RequestLogRecord(request, request_api="statistics_email_list_key")
    if request.method == "POST":
        try:
            Token=json.loads(request.body)["token"]
            Uid = UserInfo().QueryUidWithToken(Token)  # 如果登录成功后就来查询UID
            if Uid != None:  # 查到了UID
                UserOperationLogRecord(request, request_api="statistics_email_list_key", uid=Uid)  # 查询到了在计入
                Result=EmailInfo().Statistics(uid=Uid)
                return JsonResponse({'message': Result, 'code': 200, })
            else:
                return JsonResponse({'message': "小宝贝这是非法查询哦(๑•̀ㅂ•́)و✧", 'code': 403, })
        except Exception as e:
            ErrorLog().Write("Web_Email_EmailList_Statistics(def)", e)
            return JsonResponse({'message': "出错了请看报错日志(๑•̀ㅂ•́)و✧", 'code': 169, })
    else:
        return JsonResponse({'message': '请使用Post请求', 'code': 500, })

"""query_email_list_key
{

	"token": "xxxx",
	"number_of_pages": "1"
}
"""
def QueryKey(request):  # 查询邮件Key
    RequestLogRecord(request, request_api="query_email_list_key")
    if request.method == "POST":
        try:
            Token=json.loads(request.body)["token"]
            NumberOfPages = json.loads(request.body)["number_of_pages"]
            Uid = UserInfo().QueryUidWithToken(Token)  # 如果登录成功后就来查询UID
            if Uid != None:  # 查到了UID
                UserOperationLogRecord(request, request_api="query_email_list_key", uid=Uid)  # 查询到了在计入
                Result=EmailInfo().ProjectQuery(uid=Uid,number_of_pages=int(NumberOfPages))
                return JsonResponse({'message': Result, 'code': 200, })
            else:
                return JsonResponse({'message': "小宝贝这是非法查询哦(๑•̀ㅂ•́)و✧", 'code': 403, })
        except Exception as e:
            ErrorLog().Write("Web_Email_EmailList_QueryKey(def)", e)
            return JsonResponse({'message': "出错了请看报错日志(๑•̀ㅂ•́)و✧", 'code': 169, })
    else:
        return JsonResponse({'message': '请使用Post请求', 'code': 500, })

"""query_email_list
{

	"token": "xxxx",
	"project_key": "xxxx"
}
"""
def Query(request):  # 查询邮件全量的数据
    RequestLogRecord(request, request_api="query_email_list")
    if request.method == "POST":
        try:
            Token=json.loads(request.body)["token"]
            ProjectKey = json.loads(request.body)["project_key"]
            Uid = UserInfo().QueryUidWithToken(Token)  # 如果登录成功后就来查询UID
            if Uid != None:  # 查到了UID
                UserOperationLogRecord(request, request_api="query_email_list", uid=Uid)  # 查询到了在计入
                Result=EmailInfo().Query(uid=Uid,project_key=ProjectKey)
                return JsonResponse({'message': ast.literal_eval(Result), 'code': 200, })
            else:
                return JsonResponse({'message': "小宝贝这是非法查询哦(๑•̀ㅂ•́)و✧", 'code': 403, })
        except Exception as e:
            ErrorLog().Write("Web_Email_EmailList_Query(def)", e)
            return JsonResponse({'message': "出错了请看报错日志(๑•̀ㅂ•́)و✧", 'code': 169, })
    else:
        return JsonResponse({'message': '请使用Post请求', 'code': 500, })